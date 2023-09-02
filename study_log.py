import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from datetime import date
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors


start_time=datetime.now()
date=date.today()
day=date.isoweekday()




day_dict={
1:"Monday",
2:"Tuesday",
3:"Wednesday",
4:"Thursday",
5:"Friday",
6:"Saturday",
7:"Sunday"
}
day_alpha=day_dict[day]


file="C:/Users/acer/Desktop/study logger.xlsx"
wb=load_workbook(file)
ws=wb["mar-w4"]
row=ws.max_row
col=ws.max_column


char=get_column_letter(col)

empty_cell_col_no=[]
def python():
    empty_cell_col_no=[]
    count_4_the_day=0
    for i in range(1,row):
        cell_day_value=ws.cell(i,1).value
        if cell_day_value==day_alpha:
            day_select=[i]
            for j in range(2,col-20):
                char=get_column_letter(j)
                cell_value_final=ws.cell(i,j).value
                if cell_value_final==None:
                    empty_cell_col_no=empty_cell_col_no+[j]  
                
                  
           
                 
    ws.cell(day_select[0],empty_cell_col_no[0]).value=1
    light_yellow_fill=PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
    ws.cell(day_select[0],empty_cell_col_no[0]).fill=light_yellow_fill
    
    print(f"!!! congratulations a session of 60 mins added in python to  cell location ({day_select[0]},{empty_cell_col_no[0]})")
    print("\n")
    print("\n")
    



def agriculture():
    empty_cell_col_no=[]
    for i in range(1,row):
        cell_day_value=ws.cell(i,1).value
        if cell_day_value==day_alpha:
            day_select=[i]
            for j in range(14,col-9):
                char=get_column_letter(j)
                cell_value_final=ws.cell(i,j).value
                
                if cell_value_final==None:
                    empty_cell_col_no=empty_cell_col_no+[j]           
            break    
                   
    ws.cell(day_select[0],empty_cell_col_no[0]).value=1
    light_yellow_fill=PatternFill(start_color='00FF8080',end_color='00FF8080',fill_type='solid')
    ws.cell(day_select[0],empty_cell_col_no[0]).fill=light_yellow_fill
    print(f"!!! congratulations a session of 60 mins added in vba to cell location ({day_select[0]},{empty_cell_col_no[0]})")
    print("\n")
    print("\n")
    
    


def news_magazine():
    empty_cell_col_no=[]
    for i in range(1,row):
        cell_day_value=ws.cell(i,1).value
        if cell_day_value==day_alpha:
            day_select=[i]
            for j in range(25,col-1):
                char=get_column_letter(j)
                cell_value_final=ws.cell(i,j).value
                    
                if cell_value_final==None:
                    empty_cell_col_no=empty_cell_col_no+[j]
            break       
                      
                       
    ws.cell(day_select[0],empty_cell_col_no[0]).value=1
    greenfill=PatternFill(start_color='0033CCCC',end_color='0033CCCC',fill_type='solid')
    ws.cell(day_select[0],empty_cell_col_no[0]).fill=greenfill
    print(f"!!! congratulations a session of 60 mins added in news/magazine to cell location ({day_select[0]},{empty_cell_col_no[0]})")
    print("\n")
    print("\n")
    











you_are_studying=True

while you_are_studying:
    subject=int(input("\nEnter name of the subject to continue /1# for python /2# agriculture #  numeric key 3 for news/magazine"))
    if subject==1:
        now=datetime.now()
        time=now.strftime("%H:%M:%S")
        python()
        wb.save("C:/Users/acer/Desktop/study logger.xlsx")
        print(time)
        
    elif subject==2:
        now=datetime.now()
        time=now.strftime("%H:%M:%S")
        agriculture()
        wb.save("C:/Users/acer/Desktop/study logger.xlsx")
        print(time)
        
    elif subject==0:
        you_are_studying=False 
    
    else:
        now=datetime.now()
        time=now.strftime("%H:%M:%S")
        news_magazine()
        wb.save("C:/Users/acer/Desktop/study logger.xlsx")
        print(time)





    wb.save("C:/Users/acer/Desktop/study logger.xlsx")
                    


    