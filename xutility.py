import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def get_column_coumnt(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_column)

def read_data(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,columnno).value

def write_data(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)
    
    
def fill_green_color(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenfill=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rrownum,columnno).fill=greenfill
    workbook.save(file)
    
def fill_red_color(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redfill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rrownum,columnno).fill=redfill
    workbook.save(file)
    
    
    
    
    
    
    
    